import openpyxl
#TODO: change "rows=(3, 108)"(Control-F this)  to the needed rows from the first row to the last row + 1

class Person:
    def __init__(self,names,times,dinings,hall):
        self.name = names
        self.times = times
        self.dinings = dinings
        self.hall = hall
    def __str__(self):
        dineStr = ''
        for dine in self.dinings:
            dineStr += dine + ', '
        dineStr = dineStr[:-2]
        timeStr = ''
        for time in self.times:
            timeStr += str(time) + ', '
        timeStr = timeStr[:-2]
        newVariablesStr = ''
        if hasattr(self, 'newVars'):
            for newVar in self.newVars:
                newVariablesStr += " %s: %s " % (newVar[0], newVar[1])
        return "%s from %s wanted %s at %s%s" % (self.name, self.hall, dineStr, timeStr, newVariablesStr)
    def amount_of_combos(self):
        return len(self.times)*len(self.dinings)


#times must have the first letter by the # desired
#parameters are in the following order:
#workbook, sheet, tuplerow, tuplenames,  tupletimes, tupledinings, tuplehalls, tuple of tuplesnewVars
#tuples are the start col to the end col
#tuple of tuples are in the form name and row(multiple rows not allowed)
def excel_to_personlist(wbName='Fall-2016-DF4S-report.xlsx', sheetName = 'results', rows=(3, 108), nameCol=(2,4), timeCol=(14,18), dinCol=(8,14), hallCol=6, newVars=(("Email", 4), ("Phone Number", 5))):
    wb = openpyxl.load_workbook(wbName)
    sheet = wb.get_sheet_by_name(sheetName)
    personList = []
    for row in range(rows[0], rows[1]):
        newName = ''
        for col in range(nameCol[0], nameCol[1]):
            newName += sheet.cell(row=row, column = col).value + ' '
        newName = newName[:-1]
        newTimes = []
        for col in range(timeCol[0], timeCol[1]):
            x = sheet.cell(row=row, column=col).value[:1]
            #^takes the first # in the time string as the time(8:00pm -> 8)
            if x != '':
                newTimes.append(int(x))
        newDining = []
        #check for any
        if sheet.cell(row=row, column=dinCol[0]).value != '':
            for col in range(dinCol[0]+1, dinCol[1]):
                newDining.append(sheet.cell(row=rows[0]-1, column=col).value)
                #cheese that assumes the titles of the rest of the columns are the dining hall names
        else:
            for col in range(dinCol[0]+1, dinCol[1]):
                x = sheet.cell(row=row, column=col).value
                if x != '':
                    newDining.append(x)
        newHall = sheet.cell(row=row, column=hallCol).value
        newPerson = Person(newName, newTimes, newDining, newHall)
        newPerson.newVars = []
        for newVar in newVars:
            newPerson.newVars.append((newVar[0], sheet.cell(row=row, column=newVar[1]).value))
        personList.append(newPerson)
    return personList
